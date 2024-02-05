import os
import sys

from win32com import client

from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import Qt
from plyer import notification
from openpyxl import *

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

os.system('cls')

class Main(QMainWindow, QWidget):
    def __init__(self):
        super().__init__()
        self.init()
        self.site()
        self.show()

    def init(self):
        # Main window basics.
        self.setWindowIcon(QIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon)))
        self.setWindowTitle('DeskPy - XL-PDF drawer')
        self.setMinimumWidth(1200)
        self.setMinimumHeight(600)

    def site(self):
        # Login page.
        self.widget_login = QWidget()
        self.layout_p1 = QVBoxLayout()
        self.layout_p1.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignHCenter)
        self.widget_login.setLayout(self.layout_p1)

        # UI items.
        p1_h1 = QLabel('XL-PDF drawer')
        p1_h1.setObjectName('p1-h1')
        p1_h1.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        p1_h2 = QLabel('Visit us on')
        p1_h2.setObjectName('p1-h2')
        p1_h2.setAlignment(Qt.AlignmentFlag.AlignRight)
        p1_visit_us = QLabel('DeskPyLab â†—')
        p1_visit_us.setObjectName('p1-visit-us')
        p1_visit_us.setCursor(Qt.CursorShape.PointingHandCursor)
        p1_visit_us.setAlignment(Qt.AlignmentFlag.AlignLeft)
        hbox_1 = QHBoxLayout()
        hbox_1.addWidget(p1_h2)
        hbox_1.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        hbox_1.addWidget(p1_visit_us)

        self.cl_label = QLabel('CODE IS REQUIRED')
        self.cl_label.setObjectName('cl-label')
        self.cl_label.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.code_listener = QLineEdit()
        self.code_listener.setObjectName('code-listener')
        self.code_listener.setMaximumWidth(300)
        self.code_listener.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        self.code_listener.setEchoMode(QLineEdit.EchoMode.Password)

        self.get_logged = QPushButton('Ingresar')
        self.get_logged.setObjectName('get-logged-btn')
        self.get_logged.setFixedWidth(300)
        self.get_logged.setCursor(Qt.CursorShape.PointingHandCursor)
        self.get_logged.setShortcut('Return')
        self.get_logged.clicked.connect(self.user_auth)

        self.layout_p1.addWidget(p1_h1)
        self.layout_p1.addLayout(hbox_1)
        self.layout_p1.addWidget(self.cl_label)
        self.layout_p1.addWidget(self.code_listener)
        self.layout_p1.addWidget(self.get_logged)

        # Logged page.
        self.widget_logged = QWidget()
        self.layout_p2 = QVBoxLayout()
        self.layout_p2.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignCenter)
        self.widget_logged.setLayout(self.layout_p2)

        # UI items.
        p2_h1 = QLabel('XL-PDF drawer')
        p2_h1.setObjectName('p2-h1')
        p2_h1.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        p2_h2 = QLabel('Visit us on')
        p2_h2.setObjectName('p2-h2')
        p2_h2.setAlignment(Qt.AlignmentFlag.AlignRight)
        p2_visit_us = QLabel('DeskPyLab â†—')
        p2_visit_us.setObjectName('p2-visit-us')
        p2_visit_us.setCursor(Qt.CursorShape.PointingHandCursor)
        p2_visit_us.setAlignment(Qt.AlignmentFlag.AlignLeft)
        hbox_2 = QHBoxLayout()
        hbox_2.addWidget(p2_h2)
        hbox_2.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        hbox_2.addWidget(p2_visit_us)

        self.layout_p2.addWidget(p2_h1)
        self.layout_p2.addLayout(hbox_2)

        self.step_1()

        self.stackedlayout = QStackedLayout()
        self.stackedlayout.addWidget(self.widget_login)
        self.stackedlayout.addWidget(self.widget_logged)
        self.stackedlayout.setCurrentIndex(0)

        self.centralwidget = QWidget()
        self.centralwidget.setLayout(self.stackedlayout)
        self.setCentralWidget(self.centralwidget)

        # Auto-login.
        self.code_listener.setText('159357')
        self.get_logged.click()

    def user_auth(self):
        code_listener = self.code_listener.text()
        if code_listener == '159357': self.stackedlayout.setCurrentIndex(1)

    def step_1(self):
        try: self.ws1.deleteLater()
        except: pass

        self.ws1 = QWidget()
        self.ws1_lyt = QVBoxLayout()
        self.ws1.setLayout(self.ws1_lyt)
        self.ws1_lyt.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.light_1 = QLabel('ðŸ”´')
        labels1 = QLabel('Paso 1: Carga el libro con el estilo que deseas imprimir')
        self.btn_step_1 = QPushButton('aquÃ­')
        self.btn_step_1.setObjectName('btn-step-1')
        self.btn_step_1.clicked.connect(self.load_styles)
        self.btn_step_1.setCursor(Qt.CursorShape.PointingHandCursor)

        wrappers1 = QHBoxLayout()
        wrappers1.addWidget(self.light_1)
        wrappers1.addWidget(labels1)
        wrappers1.addWidget(self.btn_step_1)

        self.ws1_lyt.addLayout(wrappers1)
        self.layout_p2.addWidget(self.ws1)

    def load_styles(self):
        try: self.style_options.deleteLater()
        except: pass

        self.style_sheet = QFileDialog().getOpenFileName(filter='Excel (*.xlsx)')[0]

        if self.style_sheet != '':
            wb = load_workbook(self.style_sheet)
            sheets = wb.sheetnames

            self.style_options = QComboBox()
            self.style_options.setPlaceholderText('Select an style sheet from your book')

            for sheet in sheets:
                self.style_options.addItem(sheet)

            self.style_options.setCurrentIndex(0)
            self.ws1_lyt.addWidget(self.style_options)
            self.light_1.setText('ðŸŸ¢')

            self.step_2()

        else: self.light_1.setText('ðŸ”´')

    def step_2(self):
        try: self.ws2.deleteLater()
        except: pass

        self.ws2 = QWidget()
        self.ws2_lyt = QVBoxLayout()
        self.ws2.setLayout(self.ws2_lyt)
        self.ws2_lyt.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.light_2 = QLabel('ðŸ”´')
        labels2 = QLabel('Paso 2: Carga una lista de datos')
        self.btn_step_2 = QPushButton('aquÃ­')
        self.btn_step_2.setObjectName('btn-step-2')
        self.btn_step_2.clicked.connect(self.pull_dinamic_data)
        self.btn_step_2.setCursor(Qt.CursorShape.PointingHandCursor)

        wrappers2 = QHBoxLayout()
        wrappers2.addWidget(self.light_2)
        wrappers2.addWidget(labels2)
        wrappers2.addWidget(self.btn_step_2)

        self.ws2_lyt.addLayout(wrappers2)
        self.layout_p2.addWidget(self.ws2)

    def pull_dinamic_data(self):
        try: self.loaded_data.deleteLater()
        except: pass

        self.raw_data = QFileDialog().getOpenFileName(filter='Excel (*.xlsx)')[0]

        if self.raw_data != '':
            displayed_text_path = self.raw_data
            displayed_text_path = displayed_text_path.split('/')
            displayed_text_path = f'../{displayed_text_path[-2]}/{displayed_text_path[-1]}'
            self.loaded_data = QLabel(displayed_text_path)
            self.loaded_data.setObjectName('loaded-data')
            self.loaded_data.setAlignment(Qt.AlignmentFlag.AlignHCenter)
            self.ws2_lyt.addWidget(self.loaded_data)
            self.light_2.setText('ðŸŸ¢')

            self.build_match_system()

            self.step_3()

        else:
            try: self.loaded_data.deleteLater()
            except: pass
            self.light_2.setText('ðŸ”´')

    def build_match_system(self):
        self.header_cols = []

        # wb = load_workbook(self.loaded_data.text())
        wb = load_workbook(self.raw_data)
        ws = wb.worksheets[0]
        mc = ws.max_column

        self.record_entry_fields = []

        l = QLabel('Para cada uno de los campos abajo:')
        self.ws2_lyt.addWidget(l)
        l = QLabel('1) Indique la coordenada de la celda (correspondiente a la hoja de estilos cargada) donde se debe escribir el dato, por ejemplo: C3')
        self.ws2_lyt.addWidget(l)
        l = QLabel('2) Deje en blanco los campos de los datos que no desea usar.')
        self.ws2_lyt.addWidget(l)

        scroll = QScrollArea()
        scroll_wdg = QWidget()
        scroll_lyt = QVBoxLayout()
        scroll_wdg.setLayout(scroll_lyt)
        scroll_wdg.setMinimumHeight(200)
        scroll.setMinimumHeight(200)

        self.keep_header_meta = {}
        self.keep_header_meta = dict(self.keep_header_meta)

        for i in range(mc):
            i += 1
            self.header_cols.append(ws.cell(1,i).column_letter)

            hbx = QHBoxLayout()

            object = QLabel(f'{ws.cell(1,i).value}:')

            self.keep_header_meta[f'{ws.cell(1,i).value}'] = f'{ws.cell(1,i).column_letter}'

            hbx.addWidget(object)

            object = QLineEdit()
            object.setFixedWidth(300)
            object.setAlignment(Qt.AlignmentFlag.AlignHCenter)
            object.setPlaceholderText(f'{ws.cell(1,i).value.upper()}')
            object.setStyleSheet('padding: 5px; background: #fff; color: #000; border: 1px solid #000; border-radius: 5px;')

            self.record_entry_fields.append(object)

            hbx.addWidget(object)

            scroll_lyt.addLayout(hbx)

        scroll.setWidget(scroll_wdg)
        self.ws2_lyt.addWidget(scroll)

    def step_3(self):
        try: self.ws3.deleteLater()
        except: pass

        self.ws3 = QWidget()
        self.ws3_lyt = QVBoxLayout()
        self.ws3.setLayout(self.ws3_lyt)
        self.ws3_lyt.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.light_3 = QLabel('ðŸ”´')
        labels3 = QLabel('Paso 3: Dime dÃ³nde quieres guardar los documentos PDF?')
        self.btn_step_3 = QPushButton('Buscar')
        self.btn_step_3.setObjectName('btn-step-3')
        self.btn_step_3.clicked.connect(self.datahub)
        self.btn_step_3.setCursor(Qt.CursorShape.PointingHandCursor)

        wrappers3 = QHBoxLayout()
        wrappers3.addWidget(self.light_3)
        wrappers3.addWidget(labels3)
        wrappers3.addWidget(self.btn_step_3)
        wrappers3.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.ws3_lyt.addLayout(wrappers3)
        self.layout_p2.addWidget(self.ws3)

    def datahub(self):
        try: self.dirname.deleteLater()
        except: pass

        self.saving_dir = QFileDialog().getExistingDirectory()

        if self.saving_dir == '': self.saving_dir = 'Ruta no definida*'

        self.path = QLabel(self.saving_dir)
        self.path.setObjectName('book-path')
        self.path.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        self.ws3_lyt.addWidget(self.path)

        if self.saving_dir != 'Ruta no definida*':
            self.light_3.setText('ðŸŸ¢')
            self.step_4()
            self.step_5()

    def step_4(self):
        try: self.ws4.deleteLater()
        except: pass

        self.ws4 = QWidget()
        self.ws4_lyt = QVBoxLayout()
        self.ws4.setLayout(self.ws4_lyt)
        self.ws4_lyt.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.light_4 = QLabel('ðŸŸ¢')
        labels4 = QLabel('Paso 4: Elige el tÃ­tulo o nombre del documento como se guardarÃ¡:')

        wrappers4 = QHBoxLayout()
        wrappers4.addWidget(self.light_4)
        wrappers4.addWidget(labels4)
        wrappers4.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.ws4_lyt.addLayout(wrappers4)
        self.layout_p2.addWidget(self.ws4)

        wg = QWidget()
        ly = QHBoxLayout()
        wg.setLayout(ly)
        sc = QScrollArea()

        self.collect_all_cb = []

        for i in self.keep_header_meta:
            cbgroup = QComboBox()
            cbgroup.addItem('No usar')

            for khm in self.keep_header_meta:
                cbgroup.addItem(khm)

            self.collect_all_cb.append(cbgroup)
            ly.addWidget(cbgroup)

        sc.setWidget(wg)
        self.ws4_lyt.addWidget(sc)

    def step_5(self):
        try: self.ws5.deleteLater()
        except: pass

        self.ws5 = QWidget()
        self.ws5_lyt = QVBoxLayout()
        self.ws5.setLayout(self.ws5_lyt)
        self.ws5_lyt.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.light_5 = QLabel('ðŸŸ¢')
        labels5 = QLabel('Paso 5: Revisa la informaciÃ³n, ahora ya puedes crear tus documentos:')
        self.btn_step_5 = QPushButton('Dibujar')
        self.btn_step_5.setObjectName('btn-step-5')
        self.btn_step_5.clicked.connect(self.wizzard)
        self.btn_step_5.setCursor(Qt.CursorShape.PointingHandCursor)

        wrappers5 = QHBoxLayout()
        wrappers5.addWidget(self.light_5)
        wrappers5.addWidget(labels5)
        wrappers5.addWidget(self.btn_step_5)
        wrappers5.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.ws5_lyt.addLayout(wrappers5)
        self.layout_p2.addWidget(self.ws5)

        self.pbar = QProgressBar()
        self.pbar.setMaximumWidth(300)
        self.pbar.setValue(85)
        self.pbar.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        pbar_lyt = QHBoxLayout()
        pbar_lyt.addWidget(self.pbar)
        pbar_lyt.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.ws5_lyt.addLayout(pbar_lyt)

    def wizzard(self):
        os.system('taskkill /f /im excel.exe')

        self.wb_1 = load_workbook(self.style_sheet)
        self.wb_2 = load_workbook(self.raw_data)

        self.record_entry_fields_txt = []
        self.collect_all_cb_txt = []

        for ref in self.record_entry_fields:
            if ref.text().strip() != '': self.record_entry_fields_txt.append(ref.text().upper())
            else: self.record_entry_fields_txt.append(False)

        for ref in self.collect_all_cb:
            if ref.currentText() != 'No usar': self.collect_all_cb_txt.append(ref.currentText())

        self.file_outputname_req_coords = []

        for x in self.collect_all_cb_txt:
            if x in self.keep_header_meta: self.file_outputname_req_coords.append(self.keep_header_meta[x])

        self.ws1 = self.wb_1[self.style_options.currentText()]

        self.data_hub = []

        self.ws2 = self.wb_2.active

        for row in range(int(self.ws2.max_row)):
            row += 1
            data_block = []

            for col in range(int(self.ws2.max_column)):
                col += 1
                if str(self.ws2.cell(row,col).value).strip() != '': data_block.append(str(self.ws2.cell(row,col).value))
                else: data_block.append('')

            self.data_hub.append(data_block)

        self.data_hub.pop(0)

        y_length = len(self.record_entry_fields_txt)
        self.counter = 0

        for data_block in self.data_hub:
            os.system('taskkill /f /im excel.exe')

            self.counter += 1

            self.record = []

            for ij in range(y_length):
                x = self.record_entry_fields_txt[ij]
                if self.record_entry_fields_txt[ij] != False: self.ws1[self.record_entry_fields_txt[ij]].value = data_block[ij]

                self.record.append(data_block[ij])

            self.wb_1.save(self.style_sheet)
            self.wb_1.close()

            onlykeys = list(self.keep_header_meta.keys())

            self.selected_opts = []

            for item in self.collect_all_cb_txt:
                if item in onlykeys: self.selected_opts.append(self.record[onlykeys.index(item)])

            saving_dir = self.saving_dir.replace('/','\\')
            self.selected_opts = ' '.join(self.selected_opts)
            self.selected_opts = self.selected_opts.replace('\\','-').replace('/','-').replace(':','').replace('?','').replace('"','').replace('<','').replace('>','').replace('|','')
            self.output_pdf_name = r'{0}\{1}.pdf'.format(saving_dir,self.selected_opts)

            try: self.xlpdf()
            except Exception as e:
                QMessageBox.information(self, 'XL-PDF drawer',
                    f'\n{e}\t\n',
                    QMessageBox.StandardButton.Close, QMessageBox.StandardButton.Close)
                self.wb_2.close()

        for data_block in self.data_hub:
            for ij in range(y_length):
                x = self.record_entry_fields_txt[ij]

                if self.record_entry_fields_txt[ij] != False:
                    try: self.ws1[self.record_entry_fields_txt[ij]].value = ''
                    except Exception as e: pass

        self.wb_1.save(self.style_sheet)
        self.wb_1.close()

        self.step_6()

    def step_6(self):
        try: self.ws6.deleteLater()
        except: pass

        self.ws6 = QWidget()
        self.ws6_lyt = QVBoxLayout()
        self.ws6.setLayout(self.ws6_lyt)
        self.ws6_lyt.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.light_6 = QLabel('ðŸŸ¢')
        labels6 = QLabel('Paso 6: Los documentos estÃ¡n listos')
        self.btn_step_6 = QPushButton('aquÃ­')
        self.btn_step_6.setObjectName('btn-step-6')
        self.btn_step_6.clicked.connect(self.startf)
        self.btn_step_6.setCursor(Qt.CursorShape.PointingHandCursor)

        wrappers6 = QHBoxLayout()
        wrappers6.addWidget(self.light_6)
        wrappers6.addWidget(labels6)
        wrappers6.addWidget(self.btn_step_6)
        wrappers6.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.ws6_lyt.addLayout(wrappers6)
        self.layout_p2.addWidget(self.ws6)

    def startf(self):
        try: os.startfile(self.path.text())
        except:
            QMessageBox.information(self, 'XL-PDF drawer',
                f'La ruta:\n\n\t"{self.path.text()}"\t\t\n\nHa sido movida/eliminada o no existe.',
                QMessageBox.StandardButton.Close, QMessageBox.StandardButton.Close)

    def xlpdf(self):
        os.system('taskkill /f /im excel.exe')

        app = client.DispatchEx('Excel.Application')
        app.Interactive = False
        app.Visible = False

        wb = app.Workbooks.open(self.style_sheet)
        wb.worksheets(self.style_options.currentText()).Activate()

        wb.ActiveSheet.ExportAsFixedFormat(0, self.output_pdf_name)
        wb.Close()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet('''
        QWidget{
            background: #19002d;
            color: #fff;
            font-size: 13px;
        }
        QProgressBar{
            padding: 5px;
            background: #df00ff;
            color: #000;
            border: none;
            border-radius: 5px;
        }
        #p1-h1, #p2-h1{
            font-size: 35px;
        }
        #p1-h2, #p2-h2{
            font-size: 16px;
        }
        #p1-visit-us, #p2-visit-us{
            color: #df00ff;
            font-size: 16px;
        }
        #cl-label{
            margin-top: 22px;
            font-size: 12px;
        }
        #code-listener{
            padding: 9px;
            background: #fff;
            color: #641b70;
            font-size: 16px;
            border-radius: 20px;
        }
        #get-logged-btn{
            margin-top: 9px;
            padding: 9px;
            background: #19002d;
            color: #df00ff;
            font-size: 14px;
            border: 1px solid #df00ff;
            border-radius: 9px;
        }
        #get-logged-btn:hover{
            background: #29003d;
        }
        #get-logged-btn:focus{
            background: #12001e;
        }
        #btn-step-1, #btn-step-2, #btn-step-3, #btn-step-4, #btn-step-5, #btn-step-6{
            padding: 0;
            color: #ff0;
            border: none;
        }
        QComboBox{
            padding: 3px;
            color: #ce90ff;
            border: 1px solid #ce90ff;
            border-radius: 2px;
        }
        #loaded-data, #book-path{
            color: #ce90ff;
        }
    ''')
    win = Main()
    sys.exit(app.exec())